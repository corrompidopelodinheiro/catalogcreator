import os
import pandas as pd
from flask import Flask, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

app = Flask(__name__)

imagesize = (220,220)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/generate", methods=["POST"])
def generate():

    datafile = request.files["datafile"]
    images = request.files.getlist("images")

    data = pd.read_excel(datafile)

    data["SKU_ATUAL"] = data["SKU_ATUAL"].fillna("").astype(str).str.replace(".0","", regex=False).str.strip()
    data["SKU_ANTIGO"] = data["SKU_ANTIGO"].fillna("").astype(str).str.replace(".0","", regex=False).str.strip()
    data["DESCRICAO"] = data["DESCRICAO"].fillna("").astype(str).str.strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "catalog"

    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 22

    greenfill = PatternFill("solid", fgColor="92D050")
    yellowfill = PatternFill("solid", fgColor="FFFF00")

    titlefont = Font(size=14, bold=True)
    textfont = Font(size=11)

    centeralign = Alignment(horizontal="center", vertical="center")
    wrapalign = Alignment(wrap_text=True, vertical="top")

    thinborder = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    row = 1
    tempfiles = []

    for imgfile in images:

        filename = imgfile.filename
        sku = os.path.splitext(filename)[0].strip()

        item = data[data["SKU_ATUAL"] == sku]

        if item.empty:
            continue

        oldsku = item.iloc[0]["SKU_ANTIGO"]
        description = item.iloc[0]["DESCRICAO"]

        img = PILImage.open(imgfile.stream)
        img.thumbnail((220,220))

        temppath = f"temp_{sku}.png"
        img.save(temppath)
        tempfiles.append(temppath)

        ws.add_image(Image(temppath), f"B{row}")

        ws.merge_cells(f"D{row}:G{row}")
        c = ws[f"D{row}"]
        c.value = f"SKU ATUAL: {sku}"
        c.fill = greenfill
        c.font = titlefont
        c.alignment = centeralign
        c.border = thinborder

        ws.merge_cells(f"D{row+1}:G{row+3}")
        c = ws[f"D{row+1}"]
        c.value = f"DESCRIÇÃO:\n{description}"
        c.font = textfont
        c.alignment = wrapalign
        c.border = thinborder

        ws.merge_cells(f"D{row+4}:G{row+4}")
        c = ws[f"D{row+4}"]
        c.value = f"SKU ANTIGO: {oldsku}"
        c.fill = yellowfill
        c.font = textfont
        c.alignment = centeralign
        c.border = thinborder

        for r in range(row, row+5):
            for col in range(4,8):
                ws.cell(r,col).border = thinborder

        ws.row_dimensions[row].height = 35
        ws.row_dimensions[row+1].height = 38
        ws.row_dimensions[row+2].height = 38
        ws.row_dimensions[row+3].height = 38
        ws.row_dimensions[row+4].height = 30

        row += 12

    wb.save("catalogo.xlsx")

    for temp in tempfiles:
        if os.path.exists(temp):
            os.remove(temp)

    return send_file("catalogo.xlsx", as_attachment=True)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
